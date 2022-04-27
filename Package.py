### Libraries & Modules

# ----- BOX -----

from boxsdk import OAuth2, Client

# ----- BOX Developer Console -----

from BOX_Developer_Console import client_id, client_secret

# ----- Copy -----

from copy import copy

# ----- DateTime -----

from datetime import date

# ----- Numpy, PanDas -----

import numpy
import pandas

# ----- OpenPyXl -----

from openpyxl import load_workbook, Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN
from openpyxl.styles import colors, Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.styles.protection import Protection
from openpyxl.utils.cell import column_index_from_string, get_column_letter

# ----- OS -----

import os

# ----- Pillow -----

from PIL import Image, ImageTk

# ----- Re -----

import re

# ----- Selenium -----

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium import webdriver

# ----- TKInter -----

import tkinter
import tkinter.font
from tkinter.messagebox import showinfo
from tkinter import ttk
from ttkbootstrap import Style

### BOX Application

# ----- Interacting with BOX -----

class BOXApplication(Client):

    def __init__(self, OAuth2):
        super().__init__(OAuth2)

    def __str__(self):
        return f'{super().__str__()}'

    # ----- File ID -----

    def get_file_id(self, file_name, folder_id):

        file_id = None

        file_name_with_extension = file_name.split('.')
        file_name = file_name_with_extension[0]
        file_extension = file_name_with_extension[1]

        my_list = self.search().query(query = f'"{file_name}"', extension = [f'{file_extension}'], limit = 1)

        for i in my_list:
            file_id = i.id
        if not file_id:
            raise Exception(f'File was not found in Folder ID: {folder_id}.')
        return file_id

    # ----- Shared Link -----

    def create_shared_link(self, file_id):

        my_shared_link = None

        try:
            my_shared_link = self.file(file_id).get_shared_link(access = 'open', allow_download = True)
            print(f"Shared Link was successfully created for File ID '{file_id}'.")
            return my_shared_link

        except BoxAPIException as Error:
            raise Exception(f"An error ocurred while trying to create the Shared Link for File ID: '{file_id}'.")

    def get_shared_link(self, file_id):

        my_shared_link = None

        try:
            my_shared_link = self.file(file_id).get().shared_link
            print(f"The Shared Link for File ID '{file_id}' is the following: ")
            return my_shared_link['download_url']

        except BoxAPIException as Error:
            raise Exception(f"An error ocurred while trying to obtain the Shared Link for File ID '{file_id}'.")

    # ----- Upload & Download -----

    def upload_file(self, folder_id, file_path):

        file_for_upload = None

        try:
            file_for_upload = self.folder(folder_id).upload(file_path)
            print(f"File was successfully uploaded to BOX Folder '{folder_id}' with the following File ID '{file_for_upload.id}'.")

        except BoxAPIException as Error:
            raise Exception(f"An error ocurred while trying to upload the file to BOX Folder '{folder_id}'.")

    def download_file(self, file_id, file_path):

        download_path = None

        try:
            download_path = open(file_path, 'wb')
            self.file(file_id).download_to(download_path)
            print(f"File ID '{file_id}' was successfully downloaded.")

        except BoxAPIException as Error:
            raise Exception(f"An error ocurred while trying to download File ID '{file_id}'.")

    # ----- Update -----

    def update_file(self, file_id, file_path):

        file_for_upload = None

        try:
            file_for_upload = self.file(file_id).update_contents(file_path)
            print(f"A new version was successfully uploaded for File ID '{file_id}'.")

        except BoxAPIException as Error:
            raise Exception(f"An error ocurred while trying to upload a new version of File ID '{file_id}'.")

# Crate an instance of BOX Application

authorization = OAuth2(client_id = client_id,
                       client_secret = client_secret)

box_application = BOXApplication(authorization)

print(f'A BOX Application instance has been created successfully!')

### Selenium & TKInter

# ----- Interacting with Selenium -----

chrome = webdriver.Chrome(os.environ['USERPROFILE'] + r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Chrome Driver/chromedriver.exe')

chrome.get('https://account.box.com/api/oauth2/authorize?response_type=code&client_id=' + client_id)

# ----- Screen Resolution -----

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# ----- Interacting with TKInter -----

root = tkinter.Tk()
root.title("Confirmation Window")

frame = ttk.Frame(root, padding = (30, 15))
frame.pack()

# ----- Style & Theme -----

style_dictionary = {'Light': ['cosmo', 'flatly', 'journal', 'litera', 'lumen', 'minty', 'pulse', 'sandstone', 'united', 'yeti'],
                    'Dark': ['darkly', 'cyborg', 'superhero', 'solar']}
style = Style(theme = 'darkly')

# ----- Functions, Variables -----

access_code = tkinter.StringVar()

def obtain_url():

    # Part 1: Authentication URL

    url = chrome.current_url
    search = re.search('code=', url)
    beggining = search.span()
    access_code = url[beggining[1]:]

    # Part 2: Access Token, Refresh Token

    access_token, refresh_token = authorization.authenticate(access_code)
    print(f'Access Code: {access_code}')
    print(f'Access Token: {access_token}, Refresh Token: {refresh_token}')

    # Part 3: Confirmation Message

    window = tkinter.Toplevel()

    window_label = tkinter.Label(window, text = 'Connection succeeded!')
    ok_button = tkinter.Button(window, text = "OK", command = window.destroy)

    window_label.pack(fill = 'x', padx = 50, pady = 5)
    ok_button.pack(fill = 'x')

# ----- Widgets -----

image = Image.open(os.environ['USERPROFILE'] + r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Grant Access BOX.png').resize((750, 600))
photo = ImageTk.PhotoImage(image)
image_label = ttk.Label(frame,
                        text = f"After granting access to the application, click the 'Authorize' button.",
                        image = photo,
                        padding = (50,5),
                        compound = "bottom")

authorize_button = ttk.Button(frame, text = 'Authorize', padding = (50, 5), command = obtain_url)
quit_button = ttk.Button(frame, text = 'Quit', padding = (50, 5), command = root.destroy)

# ----- Layout -----

image_label.config(font = ('Segoe', 12))
image_label.pack(fill = 'both', expand = True)

authorize_button.pack(side = 'left', fill = 'x', expand = True)
quit_button.pack(side = 'left', fill = 'x', expand = True)

root.mainloop()

chrome.close()

### CAS Report - Update BOX Hyperlink

# Obtain the 'List of Entity - Batch 1.xlsx' File ID
list_of_entity_file_id = box_application.get_file_id('UAT Testing File 4.xlsx', '150340348409')

# Download the file 'List of Entity - Batch 1.xlsx'
box_application.download_file(list_of_entity_file_id,
                              os.environ['USERPROFILE'] + \
                              r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Input Files/List of Entity - Batch 1.xlsx')

# Convert the file 'List of Entity - Batch 1.xlsx' into a DataFrame
list_of_entity = pandas.read_excel(os.environ['USERPROFILE'] + \
                                   r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Input Files/List of Entity - Batch 1.xlsx',
                                   sheet_name = 'Audit',
                                   header = 0)

# Remove any possible null record
print(f'{list_of_entity.isnull().sum()} \n')

list_of_entity = list_of_entity.dropna()

print(f'{list_of_entity.isnull().sum()}')

# Obtain the BOX Folder Hyperlink of all the CAS Reports included in the file 'List of Entity - Batch 1.xlsx'
def get_backup_box_link(dataframe, serie):

    backup_box_link = []

    try:
        for i in dataframe.loc[:, serie]:
            j = i.split('/')[4]
            if len(j) <= 12:
                backup_box_link.append(j)
        return backup_box_link

    except:
        raise Exception(f"An error ocurred while trying to parse through the DataFrame '{dataframe}'.")

backup_box_link = get_backup_box_link(list_of_entity, 'Entity backup BOX Link')

# Obtain each CAS Report Name and File ID
def get_cas_report(list):

    cas_report = {}

    try:
        for i in list:
            coincidence = box_application.folder(i).get_items()
            for j in coincidence:
                if j.name.find('.xls') != -1 or j.name.find('.xlsx') != -1:
                    cas_report.update({j.name + '__' + str(i): j.id})
        return cas_report

    except:
        raise Exception(f"An error ocurred while trying to iterate.")

cas_report_dictionary = get_cas_report(backup_box_link)

# Convert the CAS Report Name and File ID into a DataFrame
def transform_cas_report(dictionary):

    dictionary_key = []
    dictionary_value = []
    root_folder = []

    try:
        for i in list(dictionary.keys()):
            substring = i.split('__')[0]
            dictionary_key.append(substring)

        for j in dictionary.values():
            dictionary_value.append(j)

        for k in list(dictionary.keys()):
            substring = k.split('__')[1]
            root_folder.append(substring)

        cas_report = pandas.DataFrame({'File Name': dictionary_key,
                                       'File ID': dictionary_value,
                                       'Root Folder': root_folder})
        return cas_report

    except:
        raise Exception(f"An error ocurred while trying to iterate.")

cas_report = transform_cas_report(cas_report_dictionary)

# Download each CAS Report
for i in cas_report.index:
    box_application.download_file(cas_report['File ID'][i], os.environ['USERPROFILE'] + \
                                  r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Input Files/' + \
                                  str(i) + ' - ' + str(cas_report['File Name'][i]))

# Count the number of CAS Reports
receptacle = []

for i in cas_report.index:
    receptacle.append(i)

print(f'Report Universe: {len(receptacle)}')

# Convert each CAS Report into a DataFrame
for i in range(len(receptacle)):
    receptacle[i] = pandas.read_excel(os.environ['USERPROFILE'] + \
                                      r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Input Files/' \
                                      + str(i) + ' - ' + str(cas_report['File Name'][i]),
                                      sheet_name = 'Requests',
                                      skiprows = 6)

# Assign each CAS Report it's corresponding BOX Root Folder

for i in range(len(receptacle)):
    receptacle[i]['BOX Root Folder'] = cas_report['Root Folder'][i]

# Create the Matching Criterion from the concatenation of 'Request ID' and 'Request'
for i in range(len(receptacle)):
    receptacle[i]['Matching Criterion with BOX - Updated'] = receptacle[i].apply(lambda x: str(x['Request ID']) + ' - ' + str(x['Request']),
                                                                                 axis = 1)
    receptacle[i]['Matching Criterion with BOX - Finalized'] = receptacle[i]['Matching Criterion with BOX - Updated']

# Remove all unwanted characters from the Matching Criterion
character = ['<', '>', ':', '"', '/', '|', '\?', '\*']

for i in range(len(receptacle)):
    for j in character:
        receptacle[i]['Matching Criterion with BOX - Finalized'] = \
        receptacle[i]['Matching Criterion with BOX - Finalized'].replace(j, '', regex = True)

# Cut the Matching Criterion so that it cannot exceed a maximum of 55 characters
for i in range(len(receptacle)):
    receptacle[i]['Matching Criterion with BOX - Finalized'] = receptacle[i]['Matching Criterion with BOX - Finalized'].apply(lambda x: \
                                                                                                                              x[0:54])

# Remove any blank space from the Matching Criterion
for i in range(len(receptacle)):
    receptacle[i]['Matching Criterion with BOX - Finalized'] = receptacle[i]['Matching Criterion with BOX - Finalized'].apply(lambda x: \
                                                                                                                             x.strip())

# Convert any upper case from the Matching Criterion
for i in range(len(receptacle)):
    receptacle[i]['Matching Criterion with BOX - Letter Case'] = receptacle[i]['Matching Criterion with BOX - Finalized'].apply(lambda x: \
                                                                                                                                x.lower())

# Manually compare a few examples of the Matching Criterion
print(receptacle[0]['Matching Criterion with BOX - Updated'][89])
print('-----VERSUS-----')
print(receptacle[0]['Matching Criterion with BOX - Finalized'][89])

print(receptacle[0]['Matching Criterion with BOX - Updated'][144])
print('-----VERSUS-----')
print(receptacle[0]['Matching Criterion with BOX - Finalized'][144])

# Obtain the 'Request Templates and Response Documents' Name and File ID
def request_and_response(list):

    request_template_and_response_document = {}

    try:
        for i in list:
            box_folder_collection = box_application.folder(i).get_items()
            for j in box_folder_collection:
                if j.name.find('.xls') == -1 and j.name.find('.xlsx') == -1 and j.name.find('.zip') == -1:
                    request_template_and_response_document[j.name + ' - ' + str(i)] = j.id
        return request_template_and_response_document

    except:
        raise Exception(f"An error ocurred while trying to iterate.")

request_template_and_response_document = request_and_response(backup_box_link)

# Obtain the BOX Hyperlink for each 'Request Template and Response Document'
def get_hyperlink(dictionary):

    hyperlink_dictionary = {}

    try:
        for i in range(len(dictionary)):
            box_folder_collection = box_application.folder(list(dictionary.values())[i]).get_items()
            for j in box_folder_collection:
                hyperlink_dictionary[j.name + '__' + list(dictionary.keys())[i].split(' - ')[1]] = j.id

        return hyperlink_dictionary

    except:
        raise Exception(f"An error ocurred while trying to iterate.")

hyperlink_dictionary = get_hyperlink(request_template_and_response_document)

# Convert each Request Template and Response Document into a DataFrame
def transform_hyperlink_dictionary(dictionary):

    dictionary_key = []
    dictionary_value = []
    root_folder = []

    try:
        for i in dictionary.keys():
            substring = i.split('__')[0]
            dictionary_key.append(substring)

        for j in dictionary.values():
            dictionary_value.append(j)

        for k in dictionary.keys():
            substring = k.split('__')[1]
            root_folder.append(substring)

        hyperlink = pandas.DataFrame({'Matching Criterion with BOX - Finalized': dictionary_key,
                                      'BOX Root Folder': root_folder,
                                      'BOX Folder ID': dictionary_value})
        return hyperlink

    except:
        raise Exception(f"An error ocurred while trying to iterate.")

hyperlink = transform_hyperlink_dictionary(hyperlink_dictionary)

# Delete any duplicated value present in the Matching Criterion

print(f'----- NUMBER OF RECORDS -----')
print(f'Before: {len(hyperlink)}')

hyperlink.drop_duplicates(subset = ['Matching Criterion with BOX - Finalized', 'BOX Root Folder'],
                          keep = 'first',
                          inplace = True)

print(f'After: {len(hyperlink)}')

# Format the BOX Folder Hyperlink
hyperlink['BOX Folder ID - Format'] = hyperlink['BOX Folder ID'].apply(lambda x: 'https://ibm.ent.box.com/folder/' + str(x))

# Convert any Upper Case from the Matching Criterion
hyperlink['Matching Criterion with BOX - Letter Case']  = hyperlink['Matching Criterion with BOX - Finalized'].apply(lambda x: x.lower())

# Remove any unwanted column
hyperlink.drop(['Matching Criterion with BOX - Finalized'], axis = 1, inplace = True)

# Merge the two sources through the formatted Matching Criterion
for i in range(len(receptacle)):
    receptacle[i] = receptacle[i].merge(hyperlink,
                                        how = 'left',
                                        on = ['Matching Criterion with BOX - Letter Case', 'BOX Root Folder'])

# Rename any wanted column and remove any unwanted one
for i in range(len(receptacle)):
    receptacle[i]['Matching Criterion with BOX'] = receptacle[i]['Matching Criterion with BOX - Finalized']
    receptacle[i]['Hyperlink to CAS Response'] = receptacle[i]['BOX Folder ID - Format']

    receptacle[i].drop(['Matching Criterion with BOX - Updated',
                        'Matching Criterion with BOX - Finalized',
                        'Matching Criterion with BOX - Letter Case',
                        'BOX Folder ID',
                        'BOX Folder ID - Format',
                        'BOX Root Folder'], axis = 1, inplace = True)

# Rename any additional column present in the CAS Reports

for i in range(len(receptacle)):
    try:
        receptacle[i].rename(columns = {'Unnamed: 28': 'Additional Column'}, inplace = True)
    except:
        pass

# Calculate the percentage of BOX Hyperlinks that were properly mapped
for i in range(len(receptacle)):
    probability = 1 - (receptacle[i]['Hyperlink to CAS Response'].isnull().sum() / receptacle[i].shape[0])
    print(f'DataFrame {i}: {probability:.0%}')

# ----- Interacting with Re -----

date_pattern = '(?P<split>\s+-\s+)(?P<day>\d+)\.(?P<month>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\.(?P<year>\d+)\.'

regular_expression = re.compile(date_pattern)
cas_report['File Name - Format'] = cas_report['File Name'].apply(lambda x: regular_expression.split(x)[0])

print(cas_report['File Name'][0])
print('----- VS -----')
print(cas_report['File Name - Format'][0])

file_extension_pattern = '(?P<file_extension>.xlsx)'

regular_expression = re.compile(file_extension_pattern)
cas_report['File Name - Format'] = cas_report['File Name - Format'].apply(lambda x: regular_expression.split(x)[0])

print(cas_report['File Name'][0])
print('----- VS -----')
print(cas_report['File Name - Format'][0])

# Download each CAS Report as an Excel File
for i in range(len(receptacle)):
    receptacle[i].to_excel(os.environ['USERPROFILE'] + r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Output Files/' + \
                           str(i) + ' - ' + str(cas_report['File Name - Format'][i]) + '.xlsx',
                           sheet_name = 'Requests',
                           index = False)
    print(f"{cas_report['File Name - Format'][i]} was downloaded successfully.")

### OpenPyXL

# ----- Interacting with OpenPyXL -----

def format_row(title, start_cell, end_cell, start_column, end_column):

    try:
        worksheet_1.merge_cells(start_cell + ':' + end_cell)
        worksheet_1[start_cell].value = title
        worksheet_1[start_cell].font = Font(name = 'Arial',
                                            size = 11,
                                            b = True,
                                            color = '00000000')

        for column in worksheet_1.iter_cols(min_row = 6,
                                            max_row = 6,
                                            min_col = column_index_from_string(start_column),
                                            max_col = column_index_from_string(end_column)):
            for cell in column:
                cell.border = Border(top = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                     bottom = Side(border_style = BORDER_MEDIUM, color = '00000000'))

            for column in worksheet_1.iter_cols(min_row = 6,
                                                max_row = 6,
                                                min_col = column_index_from_string(start_column),
                                                max_col = column_index_from_string(end_column)):
                for cell in column:
                    cell.fill = PatternFill(patternType = 'solid',
                                            fgColor = colors.Color(rgb = '00C0C0C0'))

        worksheet_1[start_cell].border = Border(left = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                                top = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                                bottom = Side(border_style = BORDER_MEDIUM, color = '00000000'))

        worksheet_1[end_cell].border = Border(right = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                              top = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                              bottom = Side(border_style = BORDER_MEDIUM, color = '00000000'))

    except:
        raise Exception(f'An error ocurred while trying to parse through the Workbook.')

for i in range(len(receptacle)):

    workbook = load_workbook(os.environ['USERPROFILE'] + \
                             r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Output Files/' \
                             + str(i) + ' - ' + str(cas_report['File Name - Format'][i])+ '.xlsx')
    worksheet_1 = workbook.worksheets[0]
    worksheet_1.insert_rows(0, 6)

    maximum_row = worksheet_1.max_row
    maximum_column = worksheet_1.max_column
    print(f'Workbook {i} \nMaximum Row: {maximum_row} \nMaximum Column: {maximum_column}')

    # ----- Worksheet -----

    for row in worksheet_1.iter_rows(min_row = 6, max_row = 1500,
                                     min_col = column_index_from_string('AC'), max_col = column_index_from_string('AZ')):
        for cell in row:
            cell.fill = PatternFill(patternType = 'solid',
                                    fgColor = colors.Color(rgb = 'FFFFFFFF'))

    for column in worksheet_1.iter_cols(min_row = 1, max_row = 5, min_col = 1, max_col = column_index_from_string('AZ')):
        for cell in column:
            cell.fill = PatternFill(patternType = 'solid',
                                    fgColor = colors.Color(rgb = 'FFFFFFFF'))

    for column in worksheet_1.iter_cols(min_row = maximum_row + 1, max_row = 1500, min_col = 1, max_col = column_index_from_string('AB')):
        for cell in column:
            cell.fill = PatternFill(patternType = 'solid',
                                    fgColor = colors.Color(rgb = 'FFFFFFFF'))

    # ----- A -----

    worksheet_1['A2'].value = f'Date: {(date.today().strftime("%B %d, %Y"))}'
    worksheet_1['A2'].font = Font(name = 'Arial',
                                  size = 9,
                                  b = True,
                                  color = '00000000')

    worksheet_1['A5'].value = 'Request Details: '
    worksheet_1['A5'].font = Font(name = 'Arial',
                                  size = 9,
                                  b = True,
                                  color = '00000000')

    # ----- A:AB -----

    worksheet_1.merge_cells('A1:AB1')
    worksheet_1['A1'].value = cas_report['File Name - Format'][i].split('.')[0]
    worksheet_1['A1'].font = Font(name = 'Georgia',
                                  size = 16,
                                  b = True,
                                  color = colors.Color(rgb = '00993300'))

    # ----- A:H -----

    format_row('General Details', 'A6', 'H6', 'A', 'H')

    # ----- I:P -----

    format_row('Dates & Metrics', 'I6', 'P6', 'I', 'P')

    # ----- Q:T -----

    format_row('Assignments', 'Q6', 'T6', 'Q', 'T')

    # ----- U:W -----

    format_row('Access Restrictions', 'U6', 'W6', 'U', 'W')

    # ----- X:AC -----

    format_row('Other Details', 'X6', 'AC6', 'X', 'AC')

    # ----- Header -----

    for column in worksheet_1.iter_cols(min_row = 7, max_row = 7,
                                        min_col = column_index_from_string('A'), max_col = maximum_column):
        for cell in column:
            cell.font = Font(name = 'Arial',
                             size = 11,
                             color = 'FFFFFFFF')
            cell.fill = PatternFill(patternType = 'solid',
                                    fgColor = colors.Color(rgb = '00993300'))

            cell.alignment = Alignment(horizontal = 'left')

    # ----- Cell -----

    for row in worksheet_1.iter_rows(min_row = 8, max_row = maximum_row,
                                     min_col = column_index_from_string('A'), max_col = maximum_column):
        for cell in row:
            cell.font = Font(name = 'Arial',
                             size = 11,
                             color = '00000000')
            cell.border = Border(left = Side(border_style = BORDER_THIN, color = '00000000'),
                                 right = Side(border_style = BORDER_THIN, color = '00000000'),
                                 top = Side(border_style = BORDER_THIN, color = '00000000'),
                                 bottom = Side(border_style = BORDER_THIN, color = '00000000'))
            cell.fill = PatternFill(patternType = 'solid',
                                    fgColor = colors.Color(rgb = 'FFFFFFFF'))

    # ----- Border -----

    for row in worksheet_1.iter_rows(min_row = 7, max_row = maximum_row,
                                     min_col = column_index_from_string('AC'), max_col = maximum_column):
        for cell in row:
            cell.border = Border(left = Side(border_style = BORDER_THIN, color = '00000000'),
                                 right = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                 top = Side(border_style = BORDER_THIN, color = '00000000'),
                                 bottom = Side(border_style = BORDER_THIN, color = '00000000'))

    for column in worksheet_1.iter_cols(min_row = maximum_row, max_row = maximum_row,
                                        min_col = 1, max_col = column_index_from_string('AC')):
        for cell in column:
            cell.border = Border(left = Side(border_style = BORDER_THIN, color = '00000000'),
                                 right = Side(border_style = BORDER_THIN, color = '00000000'),
                                 top = Side(border_style = BORDER_THIN, color = '00000000'),
                                 bottom = Side(border_style = BORDER_MEDIUM, color = '00000000'))

    worksheet_1['AC' + str(maximum_row)].border = Border(left = Side(border_style = BORDER_THIN, color = '00000000'),
                                                         right = Side(border_style = BORDER_MEDIUM, color = '00000000'),
                                                         top = Side(border_style = BORDER_THIN, color = '00000000'),
                                                         bottom = Side(border_style = BORDER_MEDIUM, color = '00000000'))

    workbook.save(os.environ['USERPROFILE'] + r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Output Files/' \
                  + str(i) + ' - ' + str(cas_report['File Name - Format'][i]) + '.xlsx')

    print(f'Workbook {i} has been successfully updated. \n')

### CAS Report - Upload to BOX Folder

for i in cas_report.index:
    box_application.update_file(cas_report['File ID'][i],
                                os.environ['USERPROFILE'] + \
                                r'/Desktop/ODA112 - Audit Backup Response - PwC Connect/Output Files/' + \
                                str(i) + ' - ' + str(cas_report['File Name - Format'][i]) + '.xlsx')

# ----- Screen Resolution -----

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# ----- Interacting with TKInter -----

root = tkinter.Tk()
root.title("End-Line Window")

frame = ttk.Frame(root, padding = (30, 15))
frame.pack()

# ----- Style & Theme -----

style_dictionary = {'Light': ['cosmo', 'flatly', 'journal', 'litera', 'lumen', 'minty', 'pulse', 'sandstone', 'united', 'yeti'],
                    'Dark': ['darkly', 'cyborg', 'superhero', 'solar']}
style = Style(theme = 'darkly')

# ----- Widgets -----

header = ttk.Label(frame, text = 'The following CAS Report should be now updated in BOX: ', padding = (10, 5))
cas_report = ttk.Label(frame, text = f"{cas_report['File Name - Format'][0:]}", padding = (10, 5))

ok_button = ttk.Button(frame, text = 'OK', padding = (100, 5), command = root.destroy)

# ----- Layout -----

header.pack(side = 'top', fill = 'x', expand = True)
header.config(font = ('Segoe', 12))
cas_report.pack(side = 'top', fill = 'x', expand = True)
cas_report.config(font = ('Segoe', 12))

ok_button.pack(side = 'top', fill = 'x', expand = True, padx = 5)

root.mainloop()

### End
